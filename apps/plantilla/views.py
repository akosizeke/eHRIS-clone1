import json
import re
from io import BytesIO
from urllib.parse import quote, urlencode
from zipfile import BadZipFile

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from apps.organization.models import Office

from .forms import ItemForm, NonPlantillaEmployeeForm, SalaryScheduleForm
from .models import (
    Item,
    NonPlantillaEmployee,
    SalaryGrade,
    SalaryGradeStep,
    SalarySchedule,
)


DEFAULT_SALARY_SCHEDULE_NAME = 'Default Salary Schedule'


def _active_salary_schedule():
    return SalarySchedule.objects.filter(
        is_active=True,
        effective_date__lte=timezone.localdate(),
    ).order_by('-effective_date', 'name').first()


def _get_or_create_active_salary_schedule():
    schedule = _active_salary_schedule()
    if schedule:
        return schedule

    schedule, _ = SalarySchedule.objects.get_or_create(
        name=DEFAULT_SALARY_SCHEDULE_NAME,
        defaults={
            'description': 'Default container for existing salary grade data.',
            'effective_date': timezone.localdate(),
            'is_active': True,
        },
    )
    return schedule


def _selected_salary_schedule(request, create=False):
    schedule_id = request.GET.get('schedule') or request.POST.get('schedule')
    if schedule_id:
        return get_object_or_404(SalarySchedule, pk=schedule_id)
    if create:
        return _get_or_create_active_salary_schedule()
    return _active_salary_schedule()


def _salary_grade_url(schedule=None, **params):
    query = {}
    if schedule:
        query['schedule'] = str(schedule.pk)
    query.update({
        key: value
        for key, value in params.items()
        if value not in (None, '')
    })
    url = reverse('plantilla:salary_grade')
    if query:
        return f'{url}?{urlencode(query)}'
    return url


# Converts a plantilla item model into JSON for API-style responses.
def _item_payload(item):
    return {
        'id': str(item.id),
        'item_number': item.item_number,
        'employee_name': item.employee_name,
        'position_title': item.position_title,
        'appointment_type': item.appointment_type,
        'salary_grade': item.salary_grade,
        'office_id': str(item.office_id),
        'employment_type': item.employment_type,
        'funding_source': item.funding_source,
        'position_status': item.position_status,
        'duties_responsibilities': item.duties_responsibilities,
        'legalbasis_id': str(item.legalbasis_id) if item.legalbasis_id else None,
        'created_at': item.created_at.isoformat(),
        'modified_at': item.modified_at.isoformat(),
    }


def _non_plantilla_payload(employee):
    return {
        'id': str(employee.id),
        'name': employee.name,
        'employee_type': employee.employee_type,
        'office_id': str(employee.office_id),
        'position_title': employee.position_title,
        'funding_source': employee.funding_source,
        'reference_number': employee.reference_number,
        'duties_responsibilities': employee.duties_responsibilities,
        'duration': employee.duration_display,
        'duration_value': employee.duration_value,
        'duration_unit': employee.duration_unit,
        'start_date': employee.start_date.isoformat(),
        'end_date': employee.end_date.isoformat() if employee.end_date else None,
        'compensation_rate': str(employee.compensation_rate) if employee.compensation_rate is not None else None,
        'rate_basis': employee.rate_basis,
        'salary_grade': employee.salary_grade,
        'salary_step': employee.salary_step,
        'service_provider': employee.service_provider,
        'consultancy_title': employee.consultancy_title,
        'contract_amount': str(employee.contract_amount) if employee.contract_amount is not None else None,
        'work_assignment': employee.work_assignment,
        'eligible_for_permanent': employee.eligible_for_permanent,
        'created_at': employee.created_at.isoformat(),
        'modified_at': employee.modified_at.isoformat(),
    }


# Converts one plantilla history record into JSON for the detail endpoint.
def _history_payload(history):
    return {
        'id': str(history.id),
        'plantilla_item_id': str(history.plantilla_item_id),
        'old_salary_grade': history.old_salary_grade,
        'new_salary_grade': history.new_salary_grade,
        'old_office_id': str(history.old_office_id) if history.old_office_id else None,
        'new_office_id': str(history.new_office_id) if history.new_office_id else None,
        'change_type': history.change_type,
        'effective_date': history.effective_date.isoformat(),
        'legalbasis_id': str(history.legalbasis_id) if history.legalbasis_id else None,
        'created_at': history.created_at.isoformat(),
        'modified_at': history.modified_at.isoformat(),
    }


# Shows the three-tab Plantilla Management System page.
def plantilla_list(request):
    active_tab = request.GET.get('tab', 'offices')
    if active_tab not in {'offices', 'plantilla', 'non-plantilla'}:
        return HttpResponseBadRequest('Invalid plantilla tab.')

    offices = Office.objects.filter(is_active=True).order_by('level_no', 'name')
    office_search = request.GET.get('office_q', '').strip()
    search = request.GET.get('q', '').strip()
    office_id = request.GET.get('office', '').strip()
    position_status = request.GET.get('status', '').strip()
    appointment_type = request.GET.get('appointment_type', '').strip()
    salary_grade = request.GET.get('sg', '').strip()
    non_plantilla_type = request.GET.get('type', '').strip()

    allowed_statuses = {value for value, _ in Item.POSITION_STATUS_CHOICES}
    allowed_appointment_types = {value for value, _ in Item.AppointmentType.choices}
    allowed_non_plantilla_types = {
        value for value, _ in NonPlantillaEmployee.EMPLOYEE_TYPE_CHOICES
    }

    if position_status and position_status not in allowed_statuses:
        return HttpResponseBadRequest('Invalid position status.')

    if appointment_type and appointment_type not in allowed_appointment_types:
        return HttpResponseBadRequest('Invalid appointment type.')

    if salary_grade and not salary_grade.isdigit():
        return HttpResponseBadRequest('Invalid salary grade.')

    if non_plantilla_type and non_plantilla_type not in allowed_non_plantilla_types:
        return HttpResponseBadRequest('Invalid non-plantilla type.')

    permanent_items = Item.objects.filter(
        employment_type='permanent',
    ).select_related(
        'office',
        'legalbasis',
    ).order_by(
        'office__level_no',
        'office__name',
        'item_number',
    )

    office_rows = _office_rows(offices, office_search, permanent_items)

    filtered_items = permanent_items
    if search:
        filtered_items = filtered_items.filter(
            Q(employee_name__icontains=search)
            | Q(position_title__icontains=search)
            | Q(item_number__icontains=search)
        )
    if office_id:
        filtered_items = filtered_items.filter(office_id=office_id)
    if position_status:
        filtered_items = filtered_items.filter(position_status=position_status)
    if appointment_type:
        filtered_items = filtered_items.filter(appointment_type=appointment_type)
    if salary_grade:
        filtered_items = filtered_items.filter(salary_grade=int(salary_grade))

    non_plantilla_employees = NonPlantillaEmployee.objects.select_related(
        'office',
    ).order_by(
        'office__level_no',
        'office__name',
        'name',
    )
    if search:
        non_plantilla_employees = non_plantilla_employees.filter(name__icontains=search)
    if non_plantilla_type:
        non_plantilla_employees = non_plantilla_employees.filter(employee_type=non_plantilla_type)
    if office_id:
        non_plantilla_employees = non_plantilla_employees.filter(office_id=office_id)

    if _request_wants_json(request):
        return JsonResponse({
            'plantilla': [_item_payload(item) for item in filtered_items],
            'non_plantilla': [
                _non_plantilla_payload(employee)
                for employee in non_plantilla_employees
            ],
        })

    active_salary_schedule = _active_salary_schedule()
    salary_grade_options = SalaryGrade.objects.none()
    if active_salary_schedule:
        salary_grade_options = SalaryGrade.objects.filter(
            schedule=active_salary_schedule,
            is_active=True,
        )

    return render(request, 'plantilla/list.html', {
        'active_tab': active_tab,
        'office_rows': office_rows,
        'items': filtered_items,
        'non_plantilla_employees': non_plantilla_employees,
        'offices': offices,
        'office_search': office_search,
        'search': search,
        'office_id': office_id,
        'position_status': position_status,
        'appointment_type': appointment_type,
        'salary_grade': salary_grade,
        'salary_grades': salary_grade_options.values_list('grade_number', flat=True).order_by('grade_number'),
        'non_plantilla_type': non_plantilla_type,
        'position_status_choices': Item.POSITION_STATUS_CHOICES,
        'appointment_type_choices': Item.AppointmentType.choices,
        'non_plantilla_type_choices': NonPlantillaEmployee.EMPLOYEE_TYPE_CHOICES,
    })


# Shows one plantilla item and its history, as HTML or JSON.
def plantilla_detail(request, pk):
    item = get_object_or_404(
        Item.objects.select_related('office', 'legalbasis').prefetch_related('history'),
        pk=pk,
    )
    payload = _item_payload(item)
    payload['history'] = [_history_payload(history) for history in item.history.all()]
    if _request_wants_json(request):
        return JsonResponse(payload)

    return render(request, 'plantilla/detail.html', {
        'item': item,
        'history': item.history.all(),
    })


# Shows the salary grade management page and handles manual salary grade entries.
def salary_grade(request):
    errors = []
    submitted = {}

    if request.method == 'POST':
        submitted = request.POST.copy()
        amount = submitted.get('amount', '').strip()

        if not amount.isdigit() or int(amount) < 1:
            errors.append('Enter a valid amount.')

        if not errors:
            try:
                with transaction.atomic():
                    active_schedule = _selected_salary_schedule(request, create=True)
                    next_grade_number, next_step_number = _next_salary_grade_step(
                        active_schedule,
                    )
                    salary_grade_item, _ = SalaryGrade.objects.get_or_create(
                        schedule=active_schedule,
                        grade_number=next_grade_number,
                    )
                    salary_step = SalaryGradeStep(
                        salary_grade=salary_grade_item,
                        step_number=next_step_number,
                        amount=int(amount),
                        source=SalaryGradeStep.SourceType.MANUAL,
                        is_editable=True,
                    )
                    salary_step.full_clean()
                    salary_step.save()

                return redirect(_salary_grade_url(active_schedule, added=1))
            except IntegrityError:
                errors.append('Unable to add amount to the next salary grade step.')
            except ValidationError as error:
                if hasattr(error, 'message_dict'):
                    for messages in error.message_dict.values():
                        errors.extend(messages)
                else:
                    errors.extend(error.messages)

    selected_grade = request.GET.get('grade', '').strip()
    if selected_grade and not selected_grade.isdigit():
        return HttpResponseBadRequest('Invalid salary grade.')

    active_schedule = _selected_salary_schedule(request)
    salary_grades = SalaryGrade.objects.none()
    if active_schedule:
        salary_grades = SalaryGrade.objects.filter(
            schedule=active_schedule,
            is_active=True,
        ).prefetch_related('steps').order_by('grade_number')
    grade_options = salary_grades
    if selected_grade:
        salary_grades = salary_grades.filter(grade_number=int(selected_grade))

    rows = []
    for salary_grade_item in salary_grades:
        steps_by_number = {
            step.step_number: step
            for step in salary_grade_item.steps.all()
        }
        rows.append({
            'grade_number': salary_grade_item.grade_number,
            'steps': [
                f"{steps_by_number[step].amount:,}"
                if step in steps_by_number and steps_by_number[step].amount
                else '-'
                for step in range(1, 9)
            ],
        })

    if active_schedule:
        next_grade_number, next_step_number = _next_salary_grade_step(active_schedule)
    else:
        next_grade_number, next_step_number = 1, 1
    notice = ''
    if request.GET.get('added') == '1':
        notice = 'Amount added.'
    elif request.GET.get('imported'):
        notice = f"Imported {request.GET.get('imported')} new step values."
        if request.GET.get('skipped'):
            notice = f"{notice} Skipped {request.GET.get('skipped')} existing step values."

    if request.GET.get('import_error'):
        errors.append(request.GET.get('import_error'))

    return render(request, 'salary_grade/salary_grade.html', {
        'errors': errors,
        'grade_options': grade_options,
        'rows': rows,
        'selected_grade': selected_grade,
        'submitted': submitted,
        'next_grade_number': next_grade_number,
        'next_step_number': next_step_number,
        'notice': notice,
        'active_schedule': active_schedule,
        'salary_schedules': SalarySchedule.objects.order_by('-effective_date', 'name'),
    })


def salary_schedule_list(request):
    if request.method == 'POST':
        form = SalaryScheduleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(f"{reverse('plantilla:salary_schedule_list')}?created=1")
    else:
        form = SalaryScheduleForm()

    schedules = SalarySchedule.objects.annotate(
        grade_count=Count('salary_grades', distinct=True),
        step_count=Count('salary_grades__steps', distinct=True),
    ).order_by('-effective_date', 'name')

    notice = ''
    if request.GET.get('created') == '1':
        notice = 'Salary schedule created.'
    elif request.GET.get('activated') == '1':
        notice = 'Salary schedule enabled.'

    return render(request, 'salary_grade/salary_schedule_list.html', {
        'form': form,
        'notice': notice,
        'schedules': schedules,
    })


@require_POST
def salary_schedule_activate(request, schedule_id):
    schedule = get_object_or_404(SalarySchedule, pk=schedule_id)
    if not schedule.is_active:
        schedule.is_active = True
        schedule.save(update_fields=['is_active', 'updated_at'])
    return redirect(f"{reverse('plantilla:salary_schedule_list')}?activated=1")


# CAMILLE CRISOSTOMO - 2026-06-08 

def salary_grade_detail(request, grade_number):
    active_schedule = _selected_salary_schedule(request)
    if not active_schedule:
        return HttpResponseBadRequest('No active salary schedule.')

    salary_grade_item = get_object_or_404(
        SalaryGrade.objects.prefetch_related('steps'),
        schedule=active_schedule,
        grade_number=grade_number,
    )
    steps_by_number = {
        step.step_number: step
        for step in salary_grade_item.steps.all()
    }
    steps = [
        {
            'step_number': step_number,
            'amount': (
                f"{steps_by_number[step_number].amount:,}"
                if step_number in steps_by_number and steps_by_number[step_number].amount
                else '-'
            ),
            'detail': (
                steps_by_number[step_number].get_source_display()
                if step_number in steps_by_number
                else '-'
            ),
            'amount_value': (
                steps_by_number[step_number].amount
                if step_number in steps_by_number
                else ''
            ),
            'is_editable': (
                step_number in steps_by_number
                and steps_by_number[step_number].is_editable
            ),
            'is_locked': (
                step_number in steps_by_number
                and steps_by_number[step_number].is_locked
            ),
        }
        for step_number in range(1, 9)
    ]

    return render(request, 'salary_grade/salary_grade_detail.html', {
        'salary_grade': salary_grade_item,
        'steps': steps,
        'active_schedule': active_schedule,
    })


@require_POST
def salary_grade_step_update(request, grade_number, step_number):
    active_schedule = _selected_salary_schedule(request)
    if not active_schedule:
        return HttpResponseBadRequest('No active salary schedule.')

    step = get_object_or_404(
        SalaryGradeStep.objects.select_related('salary_grade'),
        salary_grade__schedule=active_schedule,
        salary_grade__grade_number=grade_number,
        step_number=step_number,
    )
    if not step.is_editable:
        return HttpResponseBadRequest('Imported salary grade steps are view-only.')

    amount = request.POST.get('amount', '').strip()
    if not amount.isdigit() or int(amount) < 1:
        return HttpResponseBadRequest('Enter a valid amount.')

    step.amount = int(amount)
    step.full_clean()
    step.save(update_fields=['amount', 'updated_at'])
    return redirect(
        f"{reverse('plantilla:salary_grade_detail', kwargs={'grade_number': grade_number})}"
        f"?{urlencode({'schedule': str(active_schedule.pk)})}"
    )


def salary_grade_export(request):
    selected_grade = request.GET.get('grade', '').strip()
    if selected_grade and not selected_grade.isdigit():
        return HttpResponseBadRequest('Invalid salary grade.')

    active_schedule = _selected_salary_schedule(request)
    salary_grades = SalaryGrade.objects.none()
    if active_schedule:
        salary_grades = SalaryGrade.objects.filter(
            schedule=active_schedule,
            is_active=True,
        ).prefetch_related('steps').order_by('grade_number')
    if selected_grade:
        salary_grades = salary_grades.filter(grade_number=int(selected_grade))

    rows = []
    for salary_grade_item in salary_grades:
        steps_by_number = {
            step.step_number: step.amount
            for step in salary_grade_item.steps.all()
        }
        rows.append([
            salary_grade_item.grade_number,
            *[steps_by_number.get(step) for step in range(1, 9)],
        ])

    workbook = _salary_grade_workbook(rows)
    filename_suffix = (
        active_schedule.effective_date.isoformat()
        if active_schedule
        else timezone.localdate().isoformat()
    )
    filename = f"salary_grade_{filename_suffix}.xlsx"
    response = HttpResponse(
        workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def salary_grade_import(request):
    if request.method != 'POST':
        return redirect('plantilla:salary_grade')

    workbook = request.FILES.get('salary_file')
    if not workbook:
        return _salary_grade_import_error('Choose an Excel file to import.')

    if not workbook.name.lower().endswith('.xlsx'):
        return _salary_grade_import_error('Only .xlsx files are supported.')

    try:
        rows = _read_salary_grade_xlsx(workbook)
    except (BadZipFile, InvalidFileException, KeyError, ValueError):
        return _salary_grade_import_error('Import failed. Check the salary grade Excel format.')

    if not rows:
        return _salary_grade_import_error('No salary grade rows found in the Excel file.')

    imported_count = 0
    skipped_count = 0
    with transaction.atomic():
        active_schedule = _selected_salary_schedule(request, create=True)
        for row in rows:
            salary_grade_item, _ = SalaryGrade.objects.get_or_create(
                schedule=active_schedule,
                grade_number=row['grade_number'],
            )
            for step_number, amount in row['steps'].items():
                _, created = SalaryGradeStep.objects.get_or_create(
                    salary_grade=salary_grade_item,
                    step_number=step_number,
                    defaults={
                        'amount': amount,
                        'source': SalaryGradeStep.SourceType.IMPORTED,
                        'is_editable': False,
                    },
                )
                if created:
                    imported_count += 1
                else:
                    skipped_count += 1

    return redirect(
        _salary_grade_url(
            active_schedule,
            imported=imported_count,
            skipped=skipped_count,
        )
    )


def _salary_grade_import_error(message):
    return redirect(f"{reverse('plantilla:salary_grade')}?import_error={quote(message)}")


def _read_salary_grade_xlsx(workbook):
    workbook.seek(0)
    sheet = load_workbook(workbook, read_only=True, data_only=True).active
    rows = [
        list(row)
        for row in sheet.iter_rows(values_only=True)
        if any(value not in (None, '') for value in row)
    ]
    if not rows:
        return []

    header_index = _salary_grade_header_index(rows)
    if header_index is None:
        raise ValueError('Missing salary grade headers.')

    headers = rows[header_index]
    grade_column = _salary_grade_column(headers)
    step_columns = _salary_step_columns(headers)
    if grade_column is None or not step_columns:
        raise ValueError('Missing salary grade columns.')

    imported_rows = []
    for row in rows[header_index + 1:]:
        grade_number = _parse_salary_grade_number(_row_value(row, grade_column))
        if not grade_number:
            continue

        steps = {}
        for step_number, column_index in step_columns.items():
            amount = _parse_salary_amount(_row_value(row, column_index))
            if amount:
                steps[step_number] = amount

        if steps:
            imported_rows.append({
                'grade_number': grade_number,
                'steps': steps,
            })

    return imported_rows


def _salary_grade_header_index(rows):
    for index, row in enumerate(rows):
        headers = [_normalize_header(value) for value in row]
        if any(header in {'salarygrade', 'salary', 'grade'} for header in headers):
            if any(header.startswith('step') for header in headers):
                return index
    return None


def _salary_grade_column(headers):
    for column_index, value in enumerate(headers):
        if _normalize_header(value) in {'salarygrade', 'salary', 'grade'}:
            return column_index
    return None


def _salary_step_columns(headers):
    columns = {}
    for column_index, value in enumerate(headers):
        match = re.fullmatch(r'step([1-8])', _normalize_header(value))
        if match:
            columns[int(match.group(1))] = column_index
    return columns


def _normalize_header(value):
    return re.sub(r'[^a-z0-9]', '', str(value).strip().lower())


def _parse_salary_grade_number(value):
    match = re.search(r'\d+', str(value))
    return int(match.group(0)) if match else None


def _parse_salary_amount(value):
    cleaned = str(value).strip().replace(',', '')
    if cleaned in {'', '-'}:
        return None
    if re.fullmatch(r'\d+(\.0+)?', cleaned):
        return int(float(cleaned))
    return None


def _row_value(row, column_index):
    return row[column_index] if column_index < len(row) else ''


def _next_salary_grade_step(schedule):
    last_step = SalaryGradeStep.objects.select_related('salary_grade').order_by(
        '-salary_grade__grade_number',
        '-step_number',
    ).filter(
        salary_grade__schedule=schedule,
    ).first()

    if last_step:
        if last_step.step_number >= 8:
            return last_step.salary_grade.grade_number + 1, 1
        return last_step.salary_grade.grade_number, last_step.step_number + 1

    last_grade = SalaryGrade.objects.filter(
        schedule=schedule,
    ).order_by('-grade_number').first()
    if last_grade:
        return last_grade.grade_number, 1

    return 1, 1


def _salary_grade_workbook(rows):
    headers = ['Salary Grade', 'Step 1', 'Step 2', 'Step 3', 'Step 4', 'Step 5', 'Step 6', 'Step 7', 'Step 8']
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Salary Grade'
    sheet.append(headers)

    for row in rows:
        sheet.append(row)

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or '')) for cell in column_cells) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = max(width, 12)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


# Creates a plantilla item through the ItemForm and redirects to its detail page.
def plantilla_create(request):
    if request.method == 'POST':
        data = _request_data(request)
        data = data.copy()
        data['employment_type'] = 'permanent'
        data['funding_source'] = 'PS'

        form = ItemForm(data, use_salary_grade_controls=True)
        if form.is_valid():
            with transaction.atomic():
                item = form.save()
            if _request_wants_json(request):
                return JsonResponse(_item_payload(item), status=201)
            return redirect(f"{reverse('plantilla:list')}?tab=plantilla")

        if _request_wants_json(request):
            return JsonResponse({
                'errors': {
                    field: [str(error) for error in errors]
                    for field, errors in form.errors.items()
                }
            }, status=400)
    else:
        form = ItemForm(
            initial={'employment_type': 'permanent', 'funding_source': 'PS'},
            use_salary_grade_controls=True,
        )

    _prepare_permanent_form(form)

    return render(request, 'plantilla/create.html', {
        'form': form,
        'title': 'Add Position',
        'submit_label': 'Save Position',
    })


def plantilla_update(request, pk):
    item = get_object_or_404(Item, pk=pk, employment_type='permanent')

    if request.method == 'POST':
        data = _request_data(request)
        data = data.copy()
        data['employment_type'] = 'permanent'
        data['funding_source'] = item.funding_source
        form = ItemForm(data, instance=item, use_salary_grade_controls=True)
        if form.is_valid():
            with transaction.atomic():
                form.save()
            return redirect(f"{reverse('plantilla:list')}?tab=plantilla")
    else:
        form = ItemForm(instance=item, use_salary_grade_controls=True)

    _prepare_permanent_form(form)

    return render(request, 'plantilla/create.html', {
        'form': form,
        'title': 'Edit Position',
        'submit_label': 'Save Changes',
    })


def non_plantilla_create(request):
    if request.method == 'POST':
        form = NonPlantillaEmployeeForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                form.save()
            return redirect(f"{reverse('plantilla:list')}?tab=non-plantilla")
    else:
        form = NonPlantillaEmployeeForm()

    return render(request, 'plantilla/non_plantilla_form.html', {
        'form': form,
        'title': 'Add Non-Plantilla',
        'submit_label': 'Save Non-Plantilla',
    })


def non_plantilla_update(request, pk):
    employee = get_object_or_404(NonPlantillaEmployee, pk=pk)

    if request.method == 'POST':
        form = NonPlantillaEmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            with transaction.atomic():
                form.save()
            return redirect(f"{reverse('plantilla:list')}?tab=non-plantilla")
    else:
        form = NonPlantillaEmployeeForm(instance=employee)

    return render(request, 'plantilla/non_plantilla_form.html', {
        'form': form,
        'title': 'Edit Non-Plantilla',
        'submit_label': 'Save Changes',
    })


def _office_rows(offices, office_search, permanent_items):
    office_queryset = offices
    if office_search:
        office_queryset = office_queryset.filter(
            Q(name__icontains=office_search)
            | Q(office_code__icontains=office_search)
        )

    office_queryset = office_queryset.annotate(
        total_positions=Count(
            'plantilla_items',
            filter=Q(plantilla_items__employment_type='permanent'),
        ),
        filled_count=Count(
            'plantilla_items',
            filter=Q(
                plantilla_items__employment_type='permanent',
                plantilla_items__position_status='filled',
            ),
        ),
        vacant_count=Count(
            'plantilla_items',
            filter=Q(
                plantilla_items__employment_type='permanent',
                plantilla_items__position_status='vacant',
            ),
        ),
        abolished_count=Count(
            'plantilla_items',
            filter=Q(
                plantilla_items__employment_type='permanent',
                plantilla_items__position_status='abolished',
            ),
        ),
    )

    positions_by_office = {}
    office_ids = [office.pk for office in office_queryset]
    for item in permanent_items.filter(office_id__in=office_ids):
        positions_by_office.setdefault(item.office_id, []).append(item)

    return [
        {
            'office': office,
            'positions': positions_by_office.get(office.pk, []),
        }
        for office in office_queryset
    ]


def _prepare_permanent_form(form):
    form.fields['employment_type'].widget = form.fields['employment_type'].hidden_widget()
    form.fields['funding_source'].widget = form.fields['funding_source'].hidden_widget()


def _request_data(request):
    if request.content_type == 'application/json':
        try:
            return json.loads(request.body.decode('utf-8') or '{}')
        except json.JSONDecodeError:
            return {}
    return request.POST


# Detects whether the caller expects JSON instead of an HTML template.
def _request_wants_json(request):
    return (
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
        or 'application/json' in request.headers.get('accept', '')
        or request.content_type == 'application/json'
    )
